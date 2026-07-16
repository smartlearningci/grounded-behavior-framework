"""Migra a matriz N1 v1 para uma distribuição v2 equilibrada."""

import argparse
import os
import re
import sys
import tempfile
from collections import Counter
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


DEFAULT_SOURCE = "datasets/matrices/Capability_Matrix_N1_10000_v1.xlsx"
DEFAULT_TARGET = "datasets/matrices/Capability_Matrix_N1_10000_v2.xlsx"
PROVIDERS = ("gemini", "groq", "mistral", "openrouter", "cerebras")
SUBSKILLS = (
    "N1A_direct_fact",
    "N1B_ignore_irrelevant",
    "N1C_select_among_similar",
    "N1D_paraphrased_question",
    "N1E_longer_context",
)
MODELS = {
    "gemini": "gemini-3.1-flash-lite",
    "groq": "llama-3.3-70b-versatile",
    "mistral": "mistral-small-latest",
    "openrouter": "free-model-fallback",
    "cerebras": "provider-selected",
}


def create_matrix_v2(
    source: str | Path = DEFAULT_SOURCE,
    target: str | Path = DEFAULT_TARGET,
    overwrite: bool = False,
) -> dict[str, object]:
    """Cria uma cópia migrada apenas depois de validar todos os invariantes."""
    source_path = Path(source).resolve()
    target_path = Path(target).resolve()
    if not source_path.is_file():
        raise FileNotFoundError(f"Source matrix does not exist: {source_path}")
    if source_path == target_path:
        raise ValueError("Source and target matrices must be different")
    if target_path.exists() and not overwrite:
        raise FileExistsError(f"Target matrix already exists: {target_path}")

    workbook = load_workbook(source_path)
    temporary: Path | None = None
    try:
        matrix = workbook["Matrix"]
        headers = [cell.value for cell in matrix[1]]
        indexes = {str(header): index for index, header in enumerate(headers)}
        required = {
            "matrix_row_id", "batch_id", "chunk_id", "subskill_code",
            "examples_per_prompt", "assigned_provider", "preferred_model",
            "prompt_version", "dataset_split_target", "generation_status",
            "output_file", "notes",
        }
        missing = sorted(required - set(indexes))
        if missing:
            raise ValueError("Matrix is missing fields: " + ", ".join(missing))

        source_rows = [list(row) for row in matrix.iter_rows(min_row=2, values_only=True)]
        target_rows: list[list[object]] = []
        n1e_before = 0
        for source_position, values in enumerate(source_rows, start=1):
            subskill = str(values[indexes["subskill_code"]])
            parts = ("A", "B") if subskill == "N1E_longer_context" else (None,)
            if subskill == "N1E_longer_context":
                n1e_before += 1
            for part_index, part in enumerate(parts):
                row = list(values)
                batch_number = _batch_number(str(row[indexes["batch_id"]]))
                subskill_index = SUBSKILLS.index(subskill)
                provider_index = (batch_number - 1 + subskill_index + part_index) % len(PROVIDERS)
                provider = PROVIDERS[provider_index]
                if part is not None:
                    row[indexes["chunk_id"]] = f"{row[indexes['chunk_id']]}{part}"
                    row[indexes["examples_per_prompt"]] = 10
                    row[indexes["notes"]] = (
                        f"source_matrix_row={source_position}; split_part={part}; "
                        "reason=long_context_token_safety"
                    )
                row[indexes["assigned_provider"]] = provider
                row[indexes["preferred_model"]] = MODELS[provider]
                row[indexes["prompt_version"]] = "n1_v2"
                row[indexes["generation_status"]] = "pending"
                row[indexes["output_file"]] = ""
                for header, index in indexes.items():
                    if "review" in header.lower():
                        row[index] = ""
                target_rows.append(row)

        for matrix_row_id, row in enumerate(target_rows, start=1):
            row[indexes["matrix_row_id"]] = matrix_row_id
        report = validate_v2_rows(target_rows, indexes)
        report.update(
            {
                "source_matrix": str(source_path),
                "target_matrix": str(target_path),
                "source_row_count": len(source_rows),
                "target_row_count": len(target_rows),
                "n1e_rows_before": n1e_before,
                "n1e_rows_after": report["subskill_counts"]["N1E_longer_context"],
                "total_examples_before": sum(
                    int(row[indexes["examples_per_prompt"]]) for row in source_rows
                ),
                "total_examples_after": report["total_examples"],
                "validation_result": "valid",
                "migration_timestamp": datetime.now(timezone.utc).isoformat(),
            }
        )
        _replace_matrix_rows(matrix, target_rows)
        _update_coverage_summary(workbook, report)
        _update_instructions(workbook)
        _write_migration_report(workbook, report)

        target_path.parent.mkdir(parents=True, exist_ok=True)
        with tempfile.NamedTemporaryFile(
            dir=target_path.parent,
            prefix=f".{target_path.name}.",
            suffix=".xlsx",
            delete=False,
        ) as file:
            temporary = Path(file.name)
        workbook.save(temporary)
        os.replace(temporary, target_path)
        return report
    finally:
        workbook.close()
        if temporary is not None:
            temporary.unlink(missing_ok=True)


def validate_v2_rows(
    rows: list[list[object]], indexes: dict[str, int]
) -> dict[str, object]:
    """Confirma cardinalidade, cobertura, campos e cruzamento provider/subskill."""
    if len(rows) != 600:
        raise ValueError(f"V2 must contain 600 rows, found {len(rows)}")
    row_ids = [row[indexes["matrix_row_id"]] for row in rows]
    chunk_ids = [str(row[indexes["chunk_id"]]) for row in rows]
    if row_ids != list(range(1, 601)) or len(set(row_ids)) != 600:
        raise ValueError("V2 matrix_row_id values must be sequential and unique")
    if len(set(chunk_ids)) != 600:
        raise ValueError("V2 chunk_id values must be unique")
    required_fields = (
        "batch_id", "chunk_id", "curriculum_level", "capability_code",
        "subskill_code", "domain", "document_style", "fact_type",
        "entity_type", "context_length", "fact_position", "distractor_type",
        "question_type", "language", "assigned_provider", "preferred_model",
        "prompt_version", "dataset_split_target", "generation_status",
    )
    for position, row in enumerate(rows, start=1):
        empty = [field for field in required_fields if row[indexes[field]] in (None, "")]
        if empty:
            raise ValueError(f"V2 row {position} has empty fields: {', '.join(empty)}")
        subskill = str(row[indexes["subskill_code"]])
        examples = int(row[indexes["examples_per_prompt"]])
        expected = 10 if subskill == "N1E_longer_context" else 20
        if examples != expected:
            raise ValueError(f"V2 row {position} must request {expected} examples")

    provider_counts = Counter(str(row[indexes["assigned_provider"]]) for row in rows)
    subskill_counts = Counter(str(row[indexes["subskill_code"]]) for row in rows)
    expected_subskills = {subskill: 100 for subskill in SUBSKILLS[:-1]}
    expected_subskills[SUBSKILLS[-1]] = 200
    if provider_counts != Counter({provider: 120 for provider in PROVIDERS}):
        raise ValueError(f"Invalid provider balance: {dict(provider_counts)}")
    if subskill_counts != Counter(expected_subskills):
        raise ValueError(f"Invalid subskill balance: {dict(subskill_counts)}")
    cross_tab = {
        provider: {
            subskill: sum(
                1 for row in rows
                if row[indexes["assigned_provider"]] == provider
                and row[indexes["subskill_code"]] == subskill
            )
            for subskill in SUBSKILLS
        }
        for provider in PROVIDERS
    }
    for provider, counts in cross_tab.items():
        if any(counts[subskill] != 20 for subskill in SUBSKILLS[:-1]) or counts[SUBSKILLS[-1]] != 40:
            raise ValueError(f"Provider/subskill confounding remains for {provider}: {counts}")
    total_examples = sum(int(row[indexes["examples_per_prompt"]]) for row in rows)
    if total_examples != 10_000:
        raise ValueError(f"V2 must request 10000 examples, found {total_examples}")
    return {
        "provider_counts": dict(provider_counts),
        "subskill_counts": dict(subskill_counts),
        "provider_subskill_counts": cross_tab,
        "total_examples": total_examples,
    }


def _batch_number(batch_id: str) -> int:
    """Extrai o número determinístico do batch."""
    match = re.fullmatch(r"B(\d+)", batch_id)
    if not match:
        raise ValueError(f"Invalid batch_id for v2 rotation: {batch_id}")
    return int(match.group(1))


def _replace_matrix_rows(sheet, rows: list[list[object]]) -> None:
    """Substitui apenas o corpo da folha Matrix e preserva o cabeçalho."""
    template = [copy(cell._style) for cell in sheet[2]]
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    for values in rows:
        sheet.append(values)
        for cell, style in zip(sheet[sheet.max_row], template):
            cell._style = copy(style)


def _update_coverage_summary(workbook, report: dict[str, object]) -> None:
    """Atualiza a folha de resumo com totais e cruzamento v2."""
    sheet = workbook["Coverage Summary"]
    sheet.delete_rows(1, sheet.max_row)
    sheet.append(["Cognitive Curriculum Coverage Matrix — Resumo v2"])
    sheet.append([])
    sheet.append(["Indicador", "Valor"])
    sheet.append(["Total de batches", 100])
    sheet.append(["Total de prompts", 600])
    sheet.append(["Total de exemplos planeados", 10000])
    sheet.append(["Prompts por provider", 120])
    sheet.append(["N1E exemplos por prompt", 10])
    sheet.append(["Distribuição", "Todas as subskills em todos os providers"])
    sheet.append([])
    sheet.append(["Provider", "Prompts", "N1A", "N1B", "N1C", "N1D", "N1E"])
    cross = report["provider_subskill_counts"]
    for provider in PROVIDERS:
        counts = cross[provider]
        sheet.append([provider, 120] + [counts[subskill] for subskill in SUBSKILLS])


def _update_instructions(workbook) -> None:
    """Atualiza instruções operacionais incompatíveis com v2."""
    sheet = workbook["Instructions"]
    sheet.delete_rows(1, sheet.max_row)
    rows = [
        ["Como utilizar a matriz v2"],
        [],
        ["Regra", "Descrição"],
        ["Prompts", "600 prompts representam 10 000 exemplos."],
        ["N1A–N1D", "100 prompts por subskill, 20 exemplos por pedido."],
        ["N1E", "200 prompts, 10 exemplos por pedido para segurança de tokens."],
        ["Providers", "120 prompts por provider; todas as subskills usam os cinco providers."],
        ["Raw", "Preservar sempre o resultado raw antes da validação."],
        ["Validação", "Resultados estruturalmente inválidos podem regressar a retry_wait."],
    ]
    for row in rows:
        sheet.append(row)


def _write_migration_report(workbook, report: dict[str, object]) -> None:
    """Cria uma folha auditável com métricas e cruzamento."""
    if "Migration Report" in workbook.sheetnames:
        del workbook["Migration Report"]
    sheet = workbook.create_sheet("Migration Report")
    for key in (
        "source_matrix", "target_matrix", "source_row_count", "target_row_count",
        "n1e_rows_before", "n1e_rows_after", "total_examples_before",
        "total_examples_after", "validation_result", "migration_timestamp",
    ):
        sheet.append([key, report[key]])
    sheet.append([])
    sheet.append(["Provider", "Prompts", "N1A", "N1B", "N1C", "N1D", "N1E"])
    cross = report["provider_subskill_counts"]
    for provider in PROVIDERS:
        sheet.append(
            [provider, report["provider_counts"][provider]]
            + [cross[provider][subskill] for subskill in SUBSKILLS]
        )


def build_parser() -> argparse.ArgumentParser:
    """Define caminhos e proteção contra overwrite implícito."""
    parser = argparse.ArgumentParser(description="Cria a matriz N1 v2 equilibrada.")
    parser.add_argument("--source", default=DEFAULT_SOURCE)
    parser.add_argument("--output", default=DEFAULT_TARGET)
    parser.add_argument("--overwrite", action="store_true")
    return parser


def main(argv: list[str] | None = None) -> int:
    """Executa a migração e imprime apenas estatísticas validadas."""
    arguments = build_parser().parse_args(argv)
    try:
        report = create_matrix_v2(arguments.source, arguments.output, arguments.overwrite)
    except (FileNotFoundError, FileExistsError, OSError, ValueError) as exc:
        print(f"Erro: {exc}", file=sys.stderr)
        return 1
    print(f"Linhas v2: {report['target_row_count']}")
    print(f"Exemplos: {report['total_examples_after']}")
    print("Providers: " + ", ".join(f"{p}={report['provider_counts'][p]}" for p in PROVIDERS))
    print("Provider/subskill balance: valid")
    print(f"Matriz v2: {Path(arguments.output)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
