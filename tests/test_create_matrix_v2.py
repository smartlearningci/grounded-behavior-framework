"""Testes da migração equilibrada da matriz N1 v2."""

import hashlib
import os
from collections import Counter
from unittest.mock import patch

from openpyxl import load_workbook

from scripts.create_matrix_v2 import PROVIDERS, SUBSKILLS, create_matrix_v2


def test_matrix_v2_exact_balances_splits_ids_and_atomic_output(tmp_path) -> None:
    """Requisito: plano v2 completo. Resultado esperado: 600 prompts, 10 000 exemplos e cruzamento exato."""
    source = "datasets/matrices/Capability_Matrix_N1_10000_v1.xlsx"
    before = hashlib.sha256(open(source, "rb").read()).hexdigest()
    target = tmp_path / "matrix_v2.xlsx"
    with patch("scripts.create_matrix_v2.os.replace", wraps=os.replace) as replace:
        report = create_matrix_v2(source, target)

    assert replace.call_count == 1
    assert report["target_row_count"] == 600
    assert report["total_examples_after"] == 10_000
    assert report["provider_counts"] == {provider: 120 for provider in PROVIDERS}
    for provider in PROVIDERS:
        assert report["provider_subskill_counts"][provider] == {
            **{subskill: 20 for subskill in SUBSKILLS[:-1]},
            SUBSKILLS[-1]: 40,
        }
    assert hashlib.sha256(open(source, "rb").read()).hexdigest() == before

    workbook = load_workbook(target, read_only=True, data_only=True)
    try:
        assert "Migration Report" in workbook.sheetnames
        rows = list(workbook["Matrix"].iter_rows(values_only=True))
        headers = list(rows[0])
        indexes = {header: index for index, header in enumerate(headers)}
        data = rows[1:]
        assert len(data) == 600
        assert len({row[indexes["matrix_row_id"]] for row in data}) == 600
        assert len({row[indexes["chunk_id"]] for row in data}) == 600
        assert "B0001_C05A" in {row[indexes["chunk_id"]] for row in data}
        assert "B0001_C05B" in {row[indexes["chunk_id"]] for row in data}
        n1e = [row for row in data if row[indexes["subskill_code"]] == SUBSKILLS[-1]]
        assert len(n1e) == 200
        assert all(row[indexes["examples_per_prompt"]] == 10 for row in n1e)
        assert all("long_context_token_safety" in row[indexes["notes"]] for row in n1e)
        assert Counter(row[indexes["assigned_provider"]] for row in data) == Counter(
            {provider: 120 for provider in PROVIDERS}
        )
    finally:
        workbook.close()
