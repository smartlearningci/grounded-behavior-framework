"""Lê e valida a folha Matrix de um workbook Excel."""

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import MatrixRow


REQUIRED_HEADERS = tuple(MatrixRow.__dataclass_fields__)
INTEGER_FIELDS = {"matrix_row_id", "examples_per_prompt"}
ALLOWED_PROVIDERS = {
    "manual",
    "gemini",
    "groq",
    "mistral",
    "openrouter",
    "cerebras",
}
ALLOWED_STATUSES = {"pending", "generated", "validated", "rejected", "review"}
ALLOWED_SPLITS = {"train", "validation", "test", "benchmark"}


def read_matrix(
    path: str | Path,
    sheet_name: str = "Matrix",
) -> list[MatrixRow]:
    """Lê linhas da folha indicada, normaliza valores e valida a matriz."""
    matrix_path = Path(path)
    if not matrix_path.is_file():
        raise FileNotFoundError(f"Matrix file does not exist: {matrix_path}")

    workbook = load_workbook(matrix_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Matrix sheet does not exist: {sheet_name}")

        worksheet = workbook[sheet_name]
        values = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(values)
        except StopIteration as exc:
            raise ValueError(f"Matrix sheet '{sheet_name}' is empty") from exc

        headers = [
            str(value).strip() if value is not None else ""
            for value in raw_headers
        ]
        missing_headers = [
            header for header in REQUIRED_HEADERS if header not in headers
        ]
        if missing_headers:
            missing = ", ".join(missing_headers)
            raise ValueError(f"Matrix is missing required headers: {missing}")

        indexes = {header: headers.index(header) for header in REQUIRED_HEADERS}
        rows: list[MatrixRow] = []
        seen_chunk_ids: set[str] = set()
        seen_row_ids: set[int] = set()

        for excel_row, raw_row in enumerate(values, start=2):
            if _is_completely_empty(raw_row):
                continue

            row_values: dict[str, Any] = {}
            for field_name, column_index in indexes.items():
                value = (
                    raw_row[column_index]
                    if column_index < len(raw_row)
                    else None
                )
                if field_name in INTEGER_FIELDS:
                    row_values[field_name] = _to_int(
                        value,
                        field_name,
                        excel_row,
                    )
                else:
                    row_values[field_name] = _to_string(value)

            row = MatrixRow(**row_values)
            if row.chunk_id in seen_chunk_ids:
                raise ValueError(
                    f"Duplicate chunk_id '{row.chunk_id}' at Excel row {excel_row}"
                )
            if row.matrix_row_id in seen_row_ids:
                raise ValueError(
                    "Duplicate matrix_row_id "
                    f"'{row.matrix_row_id}' at Excel row {excel_row}"
                )
            seen_chunk_ids.add(row.chunk_id)
            seen_row_ids.add(row.matrix_row_id)
            rows.append(row)
    finally:
        workbook.close()

    validate_matrix_rows(rows)
    return rows


def validate_matrix_rows(rows: list[MatrixRow]) -> None:
    """Valida os campos operacionais mínimos de todas as linhas."""
    if not rows:
        raise ValueError("Matrix must not be empty")

    seen_chunk_ids: set[str] = set()
    seen_row_ids: set[int] = set()
    for position, row in enumerate(rows, start=1):
        label = f"matrix row {row.matrix_row_id or position}"
        if not row.chunk_id:
            raise ValueError(f"{label}: chunk_id must not be empty")
        if not row.batch_id:
            raise ValueError(f"{label}: batch_id must not be empty")
        if not row.subskill_code:
            raise ValueError(f"{label}: subskill_code must not be empty")
        if row.examples_per_prompt <= 0:
            raise ValueError(f"{label}: examples_per_prompt must be greater than zero")
        if not row.language:
            raise ValueError(f"{label}: language must not be empty")
        if row.assigned_provider not in ALLOWED_PROVIDERS:
            raise ValueError(
                f"{label}: invalid assigned_provider '{row.assigned_provider}'"
            )
        if row.generation_status not in ALLOWED_STATUSES:
            raise ValueError(
                f"{label}: invalid generation_status '{row.generation_status}'"
            )
        if row.dataset_split_target not in ALLOWED_SPLITS:
            raise ValueError(
                f"{label}: invalid dataset_split_target "
                f"'{row.dataset_split_target}'"
            )
        if not row.prompt_version:
            raise ValueError(f"{label}: prompt_version must not be empty")
        if row.chunk_id in seen_chunk_ids:
            raise ValueError(f"Duplicate chunk_id '{row.chunk_id}'")
        if row.matrix_row_id in seen_row_ids:
            raise ValueError(f"Duplicate matrix_row_id '{row.matrix_row_id}'")
        seen_chunk_ids.add(row.chunk_id)
        seen_row_ids.add(row.matrix_row_id)


def _is_completely_empty(values: tuple[object, ...]) -> bool:
    """Indica se uma linha não contém qualquer valor significativo."""
    return all(
        value is None or (isinstance(value, str) and not value.strip())
        for value in values
    )


def _to_string(value: object) -> str:
    """Normaliza células textuais e converte células vazias em string vazia."""
    if value is None:
        return ""
    return str(value).strip()


def _to_int(value: object, field_name: str, excel_row: int) -> int:
    """Converte uma célula numérica num inteiro ou explica a linha inválida."""
    if value is None or isinstance(value, bool):
        raise ValueError(
            f"Excel row {excel_row}: {field_name} must be an integer"
        )
    try:
        number = float(str(value).strip())
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Excel row {excel_row}: {field_name} must be an integer"
        ) from exc
    if not number.is_integer():
        raise ValueError(
            f"Excel row {excel_row}: {field_name} must be an integer"
        )
    return int(number)
